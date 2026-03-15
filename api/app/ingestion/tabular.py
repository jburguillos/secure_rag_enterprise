"""Shared XLSX/Google Sheets parsing utilities."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app.config import get_settings, get_yaml_config

_HEADER_SCAN_ROWS = 10
_PREVIEW_ROW_COUNT = 3


@dataclass(frozen=True)
class TabularSheet:
    """Normalized worksheet representation used for chunk generation."""

    name: str
    index: int
    hidden: bool
    headers: list[str]
    rows: list[tuple[int, list[str]]]
    row_count: int
    column_count: int
    tabular_truncated: bool
    truncated_reason: str | None = None


@dataclass(frozen=True)
class TabularNode:
    """Serialized tabular chunk before conversion into TextNode."""

    text: str
    metadata: dict[str, Any]


@dataclass(frozen=True)
class WorkbookParseResult:
    """Parsed workbook output used by local and Drive connectors."""

    document_text: str
    sheet_map: list[dict[str, Any]]
    tabular_nodes: list[dict[str, Any]]
    warnings: list[str]
    truncated: bool


def _normalize_scalar(value: Any, *, max_chars: int) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.isoformat(sep=" ", timespec="seconds")
    elif isinstance(value, (date, time)):
        text = value.isoformat()
    else:
        text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_chars:
        return text[: max_chars - 1].rstrip() + "…"
    return text


def _trim_row(values: list[str]) -> list[str]:
    trimmed = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _row_is_textual(values: list[str]) -> bool:
    non_empty = [value for value in values if value]
    if len(non_empty) < 2:
        return False
    alpha_like = sum(1 for value in non_empty if re.search(r"[A-Za-zÀ-ÿ]", value))
    long_cells = sum(1 for value in non_empty if len(value) >= 3)
    return alpha_like >= max(1, len(non_empty) // 2) and long_cells >= max(1, len(non_empty) // 2)


def _infer_header_row(rows: list[tuple[int, list[str]]]) -> tuple[int | None, list[str]]:
    for row_idx, values in rows[:_HEADER_SCAN_ROWS]:
        trimmed = _trim_row(values)
        if trimmed and _row_is_textual(trimmed):
            return row_idx, trimmed
    return None, []


def _slug_header(text: str, fallback: str) -> str:
    slug = re.sub(r"\s+", " ", text).strip()
    return slug or fallback


def _unique_headers(raw_headers: list[str], column_count: int) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx in range(column_count):
        fallback = f"col_{get_column_letter(idx + 1)}"
        base = _slug_header(raw_headers[idx] if idx < len(raw_headers) else "", fallback)
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def _active_column_indexes(rows: list[tuple[int, list[str]]], max_columns: int) -> list[int]:
    seen: list[int] = []
    for _row_idx, values in rows:
        for idx, value in enumerate(values[:max_columns]):
            if value and idx not in seen:
                seen.append(idx)
    return seen[:max_columns]


def _preview_rows(headers: list[str], rows: list[tuple[int, list[str]]], active_indexes: list[int]) -> str:
    preview_lines: list[str] = []
    for row_idx, values in rows[:_PREVIEW_ROW_COUNT]:
        cells: list[str] = []
        for col_idx in active_indexes:
            cell = values[col_idx] if col_idx < len(values) else ""
            if not cell:
                continue
            cells.append(f"{headers[col_idx]}={cell}")
        if cells:
            preview_lines.append(f"Row {row_idx}: " + " | ".join(cells))
    return "\n".join(preview_lines)


def _sheet_summary_text(*, workbook_name: str, sheet: TabularSheet, preview: str) -> str:
    headers_preview = ", ".join(sheet.headers[: min(8, len(sheet.headers))]) or "none"
    hidden_text = "yes" if sheet.hidden else "no"
    return (
        f"Workbook: {workbook_name}\n"
        f"Sheet: {sheet.name}\n"
        f"Hidden: {hidden_text}\n"
        f"Rows: {sheet.row_count}\n"
        f"Columns: {sheet.column_count}\n"
        f"Headers: {headers_preview}\n"
        f"Preview:\n{preview or 'No preview available.'}"
    ).strip()


def _row_block_text(
    *,
    workbook_name: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
    headers: list[str],
    active_indexes: list[int],
    rows: list[tuple[int, list[str]]],
) -> tuple[str, str]:
    header_preview = ", ".join(headers[idx] for idx in active_indexes) or "none"
    row_lines: list[str] = []
    for actual_row_idx, values in rows:
        cells: list[str] = []
        for col_idx in active_indexes:
            cell = values[col_idx] if col_idx < len(values) else ""
            if not cell:
                continue
            cells.append(f"{headers[col_idx]}={cell}")
        if cells:
            row_lines.append(f"Row {actual_row_idx}: " + " | ".join(cells))
    preview = "\n".join(row_lines[:_PREVIEW_ROW_COUNT])
    return (
        f"Workbook: {workbook_name}\n"
        f"Sheet: {sheet_name}\n"
        f"Rows: {row_start}-{row_end}\n"
        f"Headers: {header_preview}\n"
        f"{chr(10).join(row_lines)}"
    ).strip(), preview


def _build_workbook_summary(workbook_name: str, sheets: list[TabularSheet]) -> str:
    lines = [
        f"Workbook: {workbook_name}",
        f"Non-empty sheets: {len(sheets)}",
    ]
    for sheet in sheets[:12]:
        hidden_text = " hidden" if sheet.hidden else ""
        lines.append(
            f"- {sheet.name}{hidden_text}: rows={sheet.row_count}, cols={sheet.column_count}, "
            f"headers={', '.join(sheet.headers[:4]) or 'none'}"
        )
    return "\n".join(lines)


def _sheet_map_entry(sheet: TabularSheet) -> dict[str, Any]:
    return {
        "sheet_name": sheet.name,
        "sheet_index": sheet.index,
        "sheet_hidden": sheet.hidden,
        "row_count": sheet.row_count,
        "column_count": sheet.column_count,
        "column_headers": sheet.headers,
        "tabular_truncated": sheet.tabular_truncated,
        "truncated_reason": sheet.truncated_reason,
    }


def _cell_range(*, active_indexes: list[int], row_start: int, row_end: int) -> str | None:
    if not active_indexes or row_start <= 0 or row_end <= 0:
        return None
    start_col = get_column_letter(active_indexes[0] + 1)
    end_col = get_column_letter(active_indexes[-1] + 1)
    return f"{start_col}{row_start}:{end_col}{row_end}"


def _collect_sheet_rows(workbook: Workbook, *, workbook_name: str) -> tuple[list[TabularSheet], list[str], bool]:
    settings = get_settings()
    cfg = get_yaml_config().get("tabular", {})
    rows_per_block = int(cfg.get("rows_per_block", settings.tabular_rows_per_block))
    max_columns = int(cfg.get("max_columns", settings.tabular_max_columns))
    max_blocks_per_sheet = int(cfg.get("max_blocks_per_sheet", settings.tabular_max_blocks_per_sheet))
    max_sheets = int(cfg.get("max_sheets_per_workbook", settings.tabular_max_sheets_per_workbook))
    max_cell_chars = int(cfg.get("max_cell_chars", settings.tabular_max_cell_chars))
    max_rows = rows_per_block * max_blocks_per_sheet + _HEADER_SCAN_ROWS

    sheets: list[TabularSheet] = []
    warnings: list[str] = []
    workbook_truncated = False

    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        if len(sheets) >= max_sheets:
            workbook_truncated = True
            warnings.append(f"workbook_truncated:max_sheets:{worksheet.title}")
            break

        collected: list[tuple[int, list[str]]] = []
        sheet_truncated = False
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [_normalize_scalar(value, max_chars=max_cell_chars) for value in list(row)[:max_columns]]
            trimmed = _trim_row(values)
            if not trimmed:
                continue
            collected.append((row_idx, trimmed))
            if len(collected) >= max_rows:
                sheet_truncated = True
                break

        if not collected:
            continue

        header_row_idx, raw_headers = _infer_header_row(collected)
        data_rows = [entry for entry in collected if header_row_idx is None or entry[0] > header_row_idx]
        active_indexes = _active_column_indexes(data_rows or collected, max_columns=max_columns)
        if not active_indexes:
            active_indexes = list(range(min(max(len(values) for _, values in collected), max_columns)))
        if not active_indexes:
            continue

        column_count = len(active_indexes)
        header_values = [
            raw_headers[idx] if header_row_idx is not None and idx < len(raw_headers) else ""
            for idx in active_indexes
        ]
        headers = _unique_headers(header_values, column_count)

        normalized_rows: list[tuple[int, list[str]]] = []
        source_rows = data_rows or collected
        for actual_row_idx, values in source_rows:
            normalized_rows.append(
                (
                    actual_row_idx,
                    [values[col_idx] if col_idx < len(values) else "" for col_idx in active_indexes],
                )
            )

        if not normalized_rows:
            continue

        truncated_reason = None
        if sheet_truncated:
            truncated_reason = (
                f"sheet '{worksheet.title}' truncated at {max_blocks_per_sheet} blocks "
                f"of {rows_per_block} rows"
            )
            warnings.append(f"sheet_truncated:{worksheet.title}")
            workbook_truncated = True

        sheets.append(
            TabularSheet(
                name=worksheet.title,
                index=sheet_index,
                hidden=worksheet.sheet_state != "visible",
                headers=headers,
                rows=normalized_rows,
                row_count=len(normalized_rows),
                column_count=column_count,
                tabular_truncated=sheet_truncated,
                truncated_reason=truncated_reason,
            )
        )

    return sheets, warnings, workbook_truncated


def _build_nodes(*, workbook_name: str, sheets: list[TabularSheet]) -> list[TabularNode]:
    settings = get_settings()
    cfg = get_yaml_config().get("tabular", {})
    nodes: list[TabularNode] = []

    workbook_summary = _build_workbook_summary(workbook_name, sheets)
    nodes.append(
        TabularNode(
            text=workbook_summary,
            metadata={
                "source_kind": "tabular",
                "tabular_node_type": "workbook_summary",
                "sheet_name": None,
                "sheet_index": None,
                "sheet_hidden": False,
                "row_start": None,
                "row_end": None,
                "cell_range": None,
                "column_headers": [],
                "table_preview": workbook_summary[:700],
            },
        )
    )

    rows_per_block = int(cfg.get("rows_per_block", settings.tabular_rows_per_block))
    max_blocks_per_sheet = int(cfg.get("max_blocks_per_sheet", settings.tabular_max_blocks_per_sheet))

    for sheet in sheets:
        preview = _preview_rows(sheet.headers, sheet.rows, list(range(len(sheet.headers))))
        sheet_summary = _sheet_summary_text(workbook_name=workbook_name, sheet=sheet, preview=preview)
        nodes.append(
            TabularNode(
                text=sheet_summary,
                metadata={
                    "source_kind": "tabular",
                    "tabular_node_type": "sheet_summary",
                    "sheet_name": sheet.name,
                    "sheet_index": sheet.index,
                    "sheet_hidden": sheet.hidden,
                    "row_start": sheet.rows[0][0] if sheet.rows else None,
                    "row_end": sheet.rows[-1][0] if sheet.rows else None,
                    "cell_range": _cell_range(
                        active_indexes=list(range(len(sheet.headers))),
                        row_start=sheet.rows[0][0] if sheet.rows else 0,
                        row_end=sheet.rows[-1][0] if sheet.rows else 0,
                    ),
                    "column_headers": sheet.headers,
                    "table_preview": preview,
                    "tabular_truncated": sheet.tabular_truncated,
                    "truncated_reason": sheet.truncated_reason,
                },
            )
        )

        for block_index in range(0, min(len(sheet.rows), rows_per_block * max_blocks_per_sheet), rows_per_block):
            block_rows = sheet.rows[block_index : block_index + rows_per_block]
            if not block_rows:
                continue
            row_start = block_rows[0][0]
            row_end = block_rows[-1][0]
            active_indexes = [
                idx
                for idx, header in enumerate(sheet.headers)
                if any(values[idx] if idx < len(values) else "" for _, values in block_rows)
            ] or list(range(len(sheet.headers)))
            block_text, preview = _row_block_text(
                workbook_name=workbook_name,
                sheet_name=sheet.name,
                row_start=row_start,
                row_end=row_end,
                headers=sheet.headers,
                active_indexes=active_indexes,
                rows=block_rows,
            )
            nodes.append(
                TabularNode(
                    text=block_text,
                    metadata={
                        "source_kind": "tabular",
                        "tabular_node_type": "row_block",
                        "sheet_name": sheet.name,
                        "sheet_index": sheet.index,
                        "sheet_hidden": sheet.hidden,
                        "row_start": row_start,
                        "row_end": row_end,
                        "cell_range": _cell_range(active_indexes=active_indexes, row_start=row_start, row_end=row_end),
                        "column_headers": [sheet.headers[idx] for idx in active_indexes],
                        "table_preview": preview,
                        "tabular_truncated": sheet.tabular_truncated,
                        "truncated_reason": sheet.truncated_reason,
                    },
                )
            )

    return nodes


def _parse_workbook(workbook: Workbook, *, workbook_name: str) -> WorkbookParseResult:
    sheets, warnings, workbook_truncated = _collect_sheet_rows(workbook, workbook_name=workbook_name)
    if not sheets:
        return WorkbookParseResult(
            document_text=f"Workbook: {workbook_name}\nNo non-empty sheets were indexed.",
            sheet_map=[],
            tabular_nodes=[],
            warnings=warnings,
            truncated=workbook_truncated,
        )

    nodes = _build_nodes(workbook_name=workbook_name, sheets=sheets)
    return WorkbookParseResult(
        document_text=nodes[0].text,
        sheet_map=[_sheet_map_entry(sheet) for sheet in sheets],
        tabular_nodes=[{"text": node.text, **node.metadata} for node in nodes],
        warnings=warnings,
        truncated=workbook_truncated,
    )


def parse_xlsx_path(path: str | Path, *, workbook_name: str | None = None) -> WorkbookParseResult:
    """Parse an XLSX workbook from disk into workbook/sheet/row-block nodes."""

    file_path = Path(path)
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    return _parse_workbook(workbook, workbook_name=workbook_name or file_path.name)


def parse_xlsx_bytes(content: bytes, *, workbook_name: str) -> WorkbookParseResult:
    """Parse an XLSX workbook from raw bytes."""

    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    return _parse_workbook(workbook, workbook_name=workbook_name)
