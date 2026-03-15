from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from app.ingestion.gdrive_connector import _download_request, load_drive_documents
from app.ingestion.local_connector import load_local_documents
from app.ingestion.parser import chunk_documents
from app.models.schemas import Citation


def _workbook_bytes(*, include_hidden: bool = False, row_count: int = 3, include_empty_sheet: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Revenue"
    sheet.append(["quarter", "region", "revenue", "margin"])
    for idx in range(1, row_count + 1):
        sheet.append([f"Q{idx}", "EMEA", f"{idx}.2M", f"{10 + idx}%"])

    if include_hidden:
        hidden = workbook.create_sheet("HiddenPipeline")
        hidden.sheet_state = "hidden"
        hidden.append(["stage", "status"])
        hidden.append(["seed", "active"])

    if include_empty_sheet:
        workbook.create_sheet("EmptySheet")

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_local_xlsx_ingestion_creates_tabular_nodes(tmp_path: Path) -> None:
    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    workbook_path = docs_dir / "pipeline_metrics.xlsx"
    workbook_path.write_bytes(_workbook_bytes())
    acl_path = tmp_path / "acl.yaml"
    acl_path.write_text("documents: {}\n", encoding="utf-8")

    docs, skipped = load_local_documents(str(docs_dir), str(acl_path))

    assert skipped == []
    assert len(docs) == 1
    metadata = docs[0].metadata or {}
    assert metadata["mimeType"] == ".xlsx"
    assert metadata["sheet_map"][0]["sheet_name"] == "Revenue"
    assert metadata["tabular_nodes"]

    nodes = chunk_documents(docs)
    assert any(node.metadata["tabular_node_type"] == "workbook_summary" for node in nodes)
    assert any(node.metadata["tabular_node_type"] == "sheet_summary" for node in nodes)
    row_blocks = [node for node in nodes if node.metadata["tabular_node_type"] == "row_block"]
    assert row_blocks
    assert row_blocks[0].metadata["sheet_name"] == "Revenue"
    assert row_blocks[0].metadata["row_start"] == 2
    assert row_blocks[0].metadata["source_kind"] == "tabular"


def test_hidden_sheet_is_indexed_and_marked_hidden(tmp_path: Path) -> None:
    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    workbook_path = docs_dir / "hidden.xlsx"
    workbook_path.write_bytes(_workbook_bytes(include_hidden=True))
    acl_path = tmp_path / "acl.yaml"
    acl_path.write_text("documents: {}\n", encoding="utf-8")

    docs, _ = load_local_documents(str(docs_dir), str(acl_path))

    metadata = docs[0].metadata or {}
    hidden_sheets = [sheet for sheet in metadata["sheet_map"] if sheet["sheet_name"] == "HiddenPipeline"]
    assert hidden_sheets
    assert hidden_sheets[0]["sheet_hidden"] is True

    nodes = chunk_documents(docs)
    hidden_nodes = [node for node in nodes if node.metadata.get("sheet_name") == "HiddenPipeline"]
    assert hidden_nodes
    assert all(node.metadata.get("sheet_hidden") is True for node in hidden_nodes)


def test_empty_sheet_is_skipped(tmp_path: Path) -> None:
    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    workbook_path = docs_dir / "empty.xlsx"
    workbook_path.write_bytes(_workbook_bytes(include_empty_sheet=True))
    acl_path = tmp_path / "acl.yaml"
    acl_path.write_text("documents: {}\n", encoding="utf-8")

    docs, _ = load_local_documents(str(docs_dir), str(acl_path))

    sheet_names = [sheet["sheet_name"] for sheet in docs[0].metadata["sheet_map"]]
    assert "Revenue" in sheet_names
    assert "EmptySheet" not in sheet_names


def test_large_sheet_is_truncated_with_warning(tmp_path: Path) -> None:
    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    workbook_path = docs_dir / "large.xlsx"
    workbook_path.write_bytes(_workbook_bytes(row_count=5105))
    acl_path = tmp_path / "acl.yaml"
    acl_path.write_text("documents: {}\n", encoding="utf-8")

    docs, _ = load_local_documents(str(docs_dir), str(acl_path))

    metadata = docs[0].metadata or {}
    assert metadata["tabular_truncated"] is True
    assert metadata["tabular_warnings"]
    assert any(sheet["tabular_truncated"] for sheet in metadata["sheet_map"])


class _FakeRequest:
    def __init__(self, response):
        self._response = response

    def execute(self):
        return self._response


class _FakeFilesResource:
    def __init__(self, listing_by_folder):
        self._listing_by_folder = listing_by_folder
        self.last_export: tuple[str, str] | None = None

    def list(self, *, q: str, **kwargs):
        folder_id = q.split("'")[1]
        return _FakeRequest({"files": self._listing_by_folder.get(folder_id, [])})

    def export_media(self, *, fileId: str, mimeType: str):
        self.last_export = (fileId, mimeType)
        return object()

    def get_media(self, *, fileId: str, supportsAllDrives: bool = True):
        return object()


class _FakeService:
    def __init__(self, listing_by_folder):
        self._files = _FakeFilesResource(listing_by_folder)

    def files(self):
        return self._files


def test_drive_xlsx_ingestion_creates_tabular_nodes(monkeypatch) -> None:
    service = _FakeService(
        {
            "root": [
                {
                    "id": "sheet-file",
                    "name": "pipeline_metrics.xlsx",
                    "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "webViewLink": "https://drive/file",
                }
            ]
        }
    )
    monkeypatch.setattr("app.ingestion.gdrive_connector._reader_instance", lambda auth_mode: object())
    monkeypatch.setattr("app.ingestion.gdrive_connector._reader_load", lambda reader, *, folder_id, file_ids, auth_mode: [])
    monkeypatch.setattr(
        "app.ingestion.gdrive_connector.download_drive_file_bytes",
        lambda **kwargs: _workbook_bytes(),
    )
    monkeypatch.setattr(
        "app.ingestion.gdrive_connector.fetch_permissions",
        lambda file_id, service: {
            "allowed_emails": [],
            "allowed_domains": [],
            "allowed_users": [],
            "allowed_groups": [],
            "is_public": False,
            "permissions_raw": [],
        },
    )

    docs, skipped = load_drive_documents(folder_id="root", auth_mode="oauth", service=service)

    assert skipped == []
    assert len(docs) == 1
    metadata = docs[0].metadata or {}
    assert metadata["mimeType"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert metadata["sheet_map"][0]["sheet_name"] == "Revenue"
    assert metadata["tabular_nodes"]


def test_google_sheets_export_to_xlsx_and_parse(monkeypatch) -> None:
    service = _FakeService({"root": []})
    request = _download_request("sheet-id", "application/vnd.google-apps.spreadsheet", service)
    assert request is not None
    assert service.files().last_export == (
        "sheet-id",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_citation_serializes_sheet_and_row_range() -> None:
    citation = Citation(
        doc_id="pipeline_metrics.xlsx",
        doc_name="pipeline_metrics.xlsx",
        node_id="node-1",
        sheet_name="Revenue",
        row_start=26,
        row_end=50,
        cell_range="A26:D50",
        tabular_node_type="row_block",
    )

    payload = citation.model_dump()
    assert payload["sheet_name"] == "Revenue"
    assert payload["row_start"] == 26
    assert payload["row_end"] == 50
    assert payload["cell_range"] == "A26:D50"
    assert payload["tabular_node_type"] == "row_block"
